import pandas as pd
import openpyxl
from datetime import datetime
from Short_Term_Brain import AgentState

# --- SETTINGS ---
MASTER_FILE = "Main Validation UK Sheet.xlsx"

def parse_date(date_val):
    """
    The Date Engine: Converts AI strings or Excel objects into Python dates
    so they can be compared mathematically.
    """
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        # Cleans up common AI date formats
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(date_val.strip(), fmt)
            except ValueError:
                continue
    return None

def enforcer_node(state: AgentState):
    """
    Agent 4: The Enforcer
    Job: Maps IDs, finds the correct 'Lease Line' by date, and performs the Surgical Slide.
    """
    l_id = state["lease_id"]
    audit_results = state["audit_results"] 
    
    print(f"🔨 Enforcer [{l_id}]: Performing Full-Header Temporal Surgery...")

    try:
        # 1. LOAD WORKBOOK
        wb = openpyxl.load_workbook(MASTER_FILE)
        ws = wb.active

        # 2. CREATE COLUMN MAP (Anchor to Row 1 Numbers)
        col_map = {str(cell.value): cell.column_letter for cell in ws[1] if cell.value}
        
        # Check if ID column exists
        if "13" not in col_map:
            return {"error_log": state["error_log"] + ["Enforcer Error: Column 13 (Lease ID) not found in Excel."]}

        # 3. IDENTIFY ALL LINES FOR THIS LEASE
        lease_rows = []
        for row in range(4, ws.max_row + 1):
            if str(ws[f"{col_map['13']}{row}"].value) == str(l_id):
                lease_rows.append({
                    "index": row,
                    "start": parse_date(ws[f"{col_map['12']}{row}"].value),
                    "end": parse_date(ws[f"{col_map['51']}{row}"].value)
                })

        if not lease_rows:
            return {"error_log": state["error_log"] + [f"Enforcer Error: Lease {l_id} not found in rows 4+."]}

        # --- THE MASTER TEMPORAL ANCHOR MAP (This is for leases, where if there's multiple lines but different payment schedules, so we MAP out whether details already exist
        # We effectivelly don't want to be repeating figures from one of headers from the map below multiple times as that how you fail audit. We want it¬
        # to say where it exactly needs to be, the correct amount of times, in the correct dates, and leave it at that. ---
        FULL_TEMPORAL_MAP = {
            "41": "40",   # Residual Value -> Residual Date
            "47": "46",   # Purchase Value -> Purchase Date
            "48": "46",   # Option Certainty -> Purchase Date
            "54": "53",   # Break Penalty -> Break Date
            "55": "53",   # Penalty Date -> Break Date
            "73": "74",   # Ext Rent -> Ext Start Date
            "75": "74",   # Ext End -> Ext Start Date
            "78": "77",   # Ext Break Penalty -> Ext Break Date
            "97": "98",   # Initial Cost -> Cost Date
            "101": "102", # Incentive -> Incentive Date
            "103": "104", # Restoration -> Restoration Date
            "145": "145", # Next CPI -> Revision Date
            "157": "157"  # Notify Date -> Notify Date
        }

        # 4. SURGICAL DATA PLACEMENT
        for field_id, value in audit_results.items():
            if field_id not in col_map or value in [None, "", "null"]: continue
            col_letter = col_map[field_id]

            # A. Global Fields: Copy to all lines of the lease
            if field_id in ["20", "35", "147", "148", "150", "158", "159", "160", "161", "163", "165"]:
                for row_data in lease_rows:
                    ws[f"{col_letter}{row_data['index']}"] = value

            # B. Surgical Slide: Use the Anchor Map to find the specific row
            elif field_id in FULL_TEMPORAL_MAP:
                anchor_col = FULL_TEMPORAL_MAP[field_id]
                event_date = parse_date(audit_results.get(anchor_col))

                if event_date:
                    placed = False
                    for row_data in lease_rows:
                        # Logic: If row covers the date, drop it there.
                        if row_data["start"] and row_data["end"]:
                            if row_data["start"] <= event_date <= row_data["end"]:
                                ws[f"{col_letter}{row_data['index']}"] = value
                                placed = True
                    
                    if not placed:
                        # Fallback: Land in the last row if date is outside known BDO lines
                        ws[f"{col_letter}{lease_rows[-1]['index']}"] = value
                else:
                    # No anchor date found? Put in last row.
                    ws[f"{col_letter}{lease_rows[-1]['index']}"] = value
            
            # C. Everything else: Default to the Last Line
            else:
                ws[f"{col_letter}{lease_rows[-1]['index']}"] = value

        # 5. COMMIT
        wb.save(MASTER_FILE)
        return {"error_log": state["error_log"] + [f"Agent 4: {l_id} Surgical Slide Complete."]}

    except Exception as e:
        return {"error_log": state["error_log"] + [f"Enforcer Error: {str(e)}"]}